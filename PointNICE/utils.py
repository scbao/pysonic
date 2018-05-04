#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author: Theo Lemaire
# @Date:   2016-09-19 22:30:46
# @Email: theo.lemaire@epfl.ch
# @Last Modified by:   Theo Lemaire
# @Last Modified time: 2018-05-04 10:49:40

""" Definition of generic utility functions used in other modules """

from enum import Enum
import os
import pickle
import tkinter as tk
from tkinter import filedialog
import inspect
import json
import yaml
from openpyxl import load_workbook
import numpy as np
import colorlog

from . import neurons


def setLogger():
    log_formatter = colorlog.ColoredFormatter(
        '%(log_color)s %(asctime)s %(message)s',
        datefmt='%d/%m/%Y %H:%M:%S:',
        reset=True,
        log_colors={
            'DEBUG': 'green',
            'INFO': 'white',
            'WARNING': 'yellow',
            'ERROR': 'red',
            'CRITICAL': 'red,bg_white',
        },
        style='%'
    )
    log_handler = colorlog.StreamHandler()
    log_handler.setFormatter(log_formatter)
    color_logger = colorlog.getLogger('PointNICE')
    color_logger.addHandler(log_handler)
    return color_logger


# Get package logger
logger = setLogger()


class InputError(Exception):
    ''' Exception raised for errors in the input. '''
    pass


class PmCompMethod(Enum):
    """ Enum: types of computation method for the intermolecular pressure """
    direct = 1
    predict = 2


def loadYAML(filepath):
    """ Load a dictionary of parameters from an external YAML file.

        :param filepath: full path to the YAML file
        :return: parameters dictionary
    """

    _, filename = os.path.split(filepath)
    logger.debug('Loading parameters from "%s"', filename)
    with open(filepath, 'r') as f:
        stream = f.read()
    params = yaml.load(stream)
    return ParseNestedDict(params)


def getLookupDir():
    """ Return the location of the directory holding lookups files.

        :return: absolute path to the directory
    """
    this_dir, _ = os.path.split(__file__)
    return this_dir + '/lookups'


def ParseNestedDict(dict_in):
    """ Loop through a nested dictionary object and convert all string fields
        to floats.
    """
    for key, value in dict_in.items():
        if isinstance(value, dict):  # If value itself is dictionary
            dict_in[key] = ParseNestedDict(value)
        elif isinstance(dict_in[key], str):
            dict_in[key] = float(dict_in[key])
    return dict_in


def OpenFilesDialog(filetype, dirname=''):
    """ Open a FileOpenDialogBox to select one or multiple file.

        The default directory and file type are given.

        :param dirname: default directory
        :param filetype: default file type
        :return: tuple of full paths to the chosen filenames
    """
    root = tk.Tk()
    root.withdraw()
    filenames = filedialog.askopenfilenames(filetypes=[(filetype + " files", '.' + filetype)],
                                            initialdir=dirname)
    if filenames:
        par_dir = os.path.abspath(os.path.join(filenames[0], os.pardir))
    else:
        par_dir = None
    return (filenames, par_dir)


def ImportExcelCol(filename, sheetname, colstr, startrow):
    """ Load a specific column of an excel workbook as a numpy array.

        :param filename: absolute or relative path to the Excel workbook
        :param sheetname: name of the Excel spreadsheet to which data is appended
        :param colstr: string of the column to import
        :param startrow: index of the first row to consider
        :return: 1D numpy array with the column data
    """

    wb = load_workbook(filename, read_only=True)
    ws = wb[sheetname]
    range_start_str = colstr + str(startrow)
    range_stop_str = colstr + str(ws.max_row)
    tmp = np.array([[i.value for i in j] for j in ws[range_start_str:range_stop_str]])
    return tmp[:, 0]


def ConstructMatrix(serialized_inputA, serialized_inputB, serialized_output):
    """ Construct a 2D output matrix from serialized input.

        :param serialized_inputA: serialized input variable A
        :param serialized_inputB: serialized input variable B
        :param serialized_output: serialized output variable
        :return: 4-tuple with vectors of unique values of A (m) and B (n),
            output variable 2D matrix (m,n) and number of holes in the matrix
    """

    As = np.unique(serialized_inputA)
    Bs = np.unique(serialized_inputB)
    nA = As.size
    nB = Bs.size

    output = np.zeros((nA, nB))
    output[:] = np.NAN
    nholes = 0
    for i in range(nA):
        iA = np.where(serialized_inputA == As[i])
        for j in range(nB):
            iB = np.where(serialized_inputB == Bs[j])
            iMatch = np.intersect1d(iA, iB)
            if iMatch.size == 0:
                nholes += 1
            elif iMatch.size > 1:
                logger.warning('Identical serialized inputs with values (%f, %f)', As[i], Bs[j])
            else:
                iMatch = iMatch[0]
                output[i, j] = serialized_output[iMatch]
    return (As, Bs, output, nholes)


def rmse(x1, x2):
    """ Compute the root mean square error between two 1D arrays """
    return np.sqrt(((x1 - x2) ** 2).mean())


def rsquared(x1, x2):
    ''' compute the R-squared coefficient between two 1D arrays '''
    residuals = x1 - x2
    ss_res = np.sum(residuals**2)
    ss_tot = np.sum((x1 - np.mean(x1))**2)
    return 1 - (ss_res / ss_tot)


def DownSample(t_dense, y, nsparse):
    """ Decimate periodic signals to a specified number of samples."""

    if(y.ndim) > 1:
        nsignals = y.shape[0]
    else:
        nsignals = 1
        y = np.array([y])

    # determine time step and period of input signal
    T = t_dense[-1] - t_dense[0]
    dt_dense = t_dense[1] - t_dense[0]

    # resample time vector linearly
    t_ds = np.linspace(t_dense[0], t_dense[-1], nsparse)

    # create MAV window
    nmav = int(0.03 * T / dt_dense)
    if nmav % 2 == 0:
        nmav += 1
    mav = np.ones(nmav) / nmav

    # determine signals padding
    npad = int((nmav - 1) / 2)

    # determine indexes of sampling on convolved signals
    ids = np.round(np.linspace(0, t_dense.size - 1, nsparse)).astype(int)

    y_ds = np.empty((nsignals, nsparse))

    # loop through signals
    for i in range(nsignals):
        # pad, convolve and resample
        pad_left = y[i, -(npad + 2):-2]
        pad_right = y[i, 1:npad + 1]
        y_ext = np.concatenate((pad_left, y[i, :], pad_right), axis=0)
        y_mav = np.convolve(y_ext, mav, mode='valid')
        y_ds[i, :] = y_mav[ids]

    if nsignals == 1:
        y_ds = y_ds[0, :]

    return (t_ds, y_ds)


def Pressure2Intensity(p, rho=1075.0, c=1515.0):
    """ Return the spatial peak, pulse average acoustic intensity (ISPPA)
        associated with the specified pressure amplitude.

        :param p: pressure amplitude (Pa)
        :param rho: medium density (kg/m3)
        :param c: speed of sound in medium (m/s)
        :return: spatial peak, pulse average acoustic intensity (W/m2)
    """
    return p**2 / (2 * rho * c)


def Intensity2Pressure(I, rho=1075.0, c=1515.0):
    """ Return the pressure amplitude associated with the specified
        spatial peak, pulse average acoustic intensity (ISPPA).

        :param I: spatial peak, pulse average acoustic intensity (W/m2)
        :param rho: medium density (kg/m3)
        :param c: speed of sound in medium (m/s)
        :return: pressure amplitude (Pa)
    """
    return np.sqrt(2 * rho * c * I)


def find_nearest(array, value):
    ''' Find nearest element in 1D array. '''

    idx = (np.abs(array - value)).argmin()
    return (idx, array[idx])


def rescale(x, lb, ub, lb_new=0, ub_new=1):
    ''' Rescale a value to a specific interval by linear transformation. '''

    xnorm = (x - lb) / (ub - lb)
    return xnorm * (ub_new - lb_new) + lb_new


def LennardJones(x, beta, alpha, C, m, n):
    """ Generic expression of a Lennard-Jones function, adapted for the context of
        symmetric deflection (distance = 2x).

        :param x: deflection (i.e. half-distance)
        :param beta: x-shifting factor
        :param alpha: x-scaling factor
        :param C: y-scaling factor
        :param m: exponent of the repulsion term
        :param n: exponent of the attraction term
        :return: Lennard-Jones potential at given distance (2x)
    """
    return C * (np.power((alpha / (2 * x + beta)), m) - np.power((alpha / (2 * x + beta)), n))


def getNeuronsDict():
    ''' Return dictionary of neurons classes that can be instantiated. '''
    neurons_dict = {}
    for _, obj in inspect.getmembers(neurons):
        if inspect.isclass(obj) and isinstance(obj.name, str):
            neurons_dict[obj.name] = obj
    return neurons_dict


def get_BLS_lookups(a):
    lookup_path = getLookupDir() + '/BLS_lookups_a{:.1f}nm.json'.format(a * 1e9)
    try:
        with open(lookup_path) as fh:
            sample = json.load(fh)
        return sample
    except FileNotFoundError:
        return {}


def save_BLS_lookups(a, lookups):
    """ Save BLS parameter into specific lookup file
        :return: absolute path to the directory
    """
    lookup_path = getLookupDir() + '/BLS_lookups_a{:.1f}nm.json'.format(a * 1e9)
    with open(lookup_path, 'w') as fh:
        json.dump(lookups, fh)


def extractCompTimes(filenames):
    ''' Extract computation times from a list of simulation files. '''
    tcomps = np.empty(len(filenames))
    for i, fn in enumerate(filenames):
        logger.info('Loading data from "%s"', fn)
        with open(fn, 'rb') as fh:
            frame = pickle.load(fh)
            meta = frame['meta']
        tcomps[i] = meta['tcomp']
    return tcomps


def computeMeshEdges(x, scale='lin'):
    ''' Compute the appropriate edges of a mesh that quads a linear or logarihtmic distribution.

        :param x: the input vector
        :param scale: the type of distribution ('lin' for linear, 'log' for logarihtmic)
        :return: the edges vector
    '''

    if scale is 'log':
        x = np.log10(x)
    dx = x[1] - x[0]
    if scale is 'lin':
        y = np.linspace(x[0] - dx / 2, x[-1] + dx / 2, x.size + 1)
    elif scale is 'log':
        y = np.logspace(x[0] - dx / 2, x[-1] + dx / 2, x.size + 1)
    return y


si_prefixes = {
    'y': 1e-24,  # yocto
    'z': 1e-21,  # zepto
    'a': 1e-18,  # atto
    'f': 1e-15,  # femto
    'p': 1e-12,  # pico
    'n': 1e-9,   # nano
    'u': 1e-6,   # micro
    'm': 1e-3,   # mili
    '': 1e0,    # None
    'k': 1e3,    # kilo
    'M': 1e6,    # mega
    'G': 1e9,    # giga
    'T': 1e12,   # tera
    'P': 1e15,   # peta
    'E': 1e18,   # exa
    'Z': 1e21,   # zetta
    'Y': 1e24,   # yotta
}


def si_format(x, precision=0):
    ''' Format a float according to the SI unit system, with the appropriate prefix letter. '''
    if isinstance(x, float) or isinstance(x, int):
        if x == 0:
            factor = 1e0
            prefix = ''
        else:
            vals = list(si_prefixes.values())
            ix = np.searchsorted(vals, np.abs(x)) - 1
            if np.abs(x) == vals[ix + 1]:
                ix += 1
            factor = vals[ix]
            prefix = list(si_prefixes.keys())[ix]
        return '{{:.{}f}}{}'.format(precision, prefix).format(x / factor)
    elif isinstance(x, list) or isinstance(x, tuple):
        return [si_format(item, precision) for item in x]
