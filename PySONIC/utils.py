#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Author: Theo Lemaire
# @Date:   2016-09-19 22:30:46
# @Email: theo.lemaire@epfl.ch
# @Last Modified by:   Theo Lemaire
# @Last Modified time: 2018-08-24 16:19:36

""" Definition of generic utility functions used in other modules """

from enum import Enum
import operator
import os
import pickle
import tkinter as tk
from tkinter import filedialog
import inspect
import json
import yaml
from openpyxl import load_workbook
import numpy as np
import colorlog

from . import neurons


def setLogger():
    log_formatter = colorlog.ColoredFormatter(
        '%(log_color)s %(asctime)s %(message)s',
        datefmt='%d/%m/%Y %H:%M:%S:',
        reset=True,
        log_colors={
            'DEBUG': 'green',
            'INFO': 'white',
            'WARNING': 'yellow',
            'ERROR': 'red',
            'CRITICAL': 'red,bg_white',
        },
        style='%'
    )
    log_handler = colorlog.StreamHandler()
    log_handler.setFormatter(log_formatter)
    color_logger = colorlog.getLogger('PySONIC')
    color_logger.addHandler(log_handler)
    return color_logger


# Get package logger
logger = setLogger()


class InputError(Exception):
    ''' Exception raised for errors in the input. '''
    pass


class PmCompMethod(Enum):
    """ Enum: types of computation method for the intermolecular pressure """
    direct = 1
    predict = 2


def loadYAML(filepath):
    """ Load a dictionary of parameters from an external YAML file.

        :param filepath: full path to the YAML file
        :return: parameters dictionary
    """

    _, filename = os.path.split(filepath)
    logger.debug('Loading parameters from "%s"', filename)
    with open(filepath, 'r') as f:
        stream = f.read()
    params = yaml.load(stream)
    return ParseNestedDict(params)


def getLookupDir():
    """ Return the location of the directory holding lookups files.

        :return: absolute path to the directory
    """
    this_dir, _ = os.path.split(__file__)
    return os.path.join(this_dir, 'lookups')


def ParseNestedDict(dict_in):
    """ Loop through a nested dictionary object and convert all string fields
        to floats.
    """
    for key, value in dict_in.items():
        if isinstance(value, dict):  # If value itself is dictionary
            dict_in[key] = ParseNestedDict(value)
        elif isinstance(dict_in[key], str):
            dict_in[key] = float(dict_in[key])
    return dict_in


def OpenFilesDialog(filetype, dirname=''):
    """ Open a FileOpenDialogBox to select one or multiple file.

        The default directory and file type are given.

        :param dirname: default directory
        :param filetype: default file type
        :return: tuple of full paths to the chosen filenames
    """
    root = tk.Tk()
    root.withdraw()
    filenames = filedialog.askopenfilenames(filetypes=[(filetype + " files", '.' + filetype)],
                                            initialdir=dirname)
    if filenames:
        par_dir = os.path.abspath(os.path.join(filenames[0], os.pardir))
    else:
        par_dir = None
    return (filenames, par_dir)


def ImportExcelCol(filename, sheetname, colstr, startrow):
    """ Load a specific column of an excel workbook as a numpy array.

        :param filename: absolute or relative path to the Excel workbook
        :param sheetname: name of the Excel spreadsheet to which data is appended
        :param colstr: string of the column to import
        :param startrow: index of the first row to consider
        :return: 1D numpy array with the column data
    """

    wb = load_workbook(filename, read_only=True)
    ws = wb[sheetname]
    range_start_str = colstr + str(startrow)
    range_stop_str = colstr + str(ws.max_row)
    tmp = np.array([[i.value for i in j] for j in ws[range_start_str:range_stop_str]])
    return tmp[:, 0]


def ConstructMatrix(serialized_inputA, serialized_inputB, serialized_output):
    """ Construct a 2D output matrix from serialized input.

        :param serialized_inputA: serialized input variable A
        :param serialized_inputB: serialized input variable B
        :param serialized_output: serialized output variable
        :return: 4-tuple with vectors of unique values of A (m) and B (n),
            output variable 2D matrix (m,n) and number of holes in the matrix
    """

    As = np.unique(serialized_inputA)
    Bs = np.unique(serialized_inputB)
    nA = As.size
    nB = Bs.size

    output = np.zeros((nA, nB))
    output[:] = np.NAN
    nholes = 0
    for i in range(nA):
        iA = np.where(serialized_inputA == As[i])
        for j in range(nB):
            iB = np.where(serialized_inputB == Bs[j])
            iMatch = np.intersect1d(iA, iB)
            if iMatch.size == 0:
                nholes += 1
            elif iMatch.size > 1:
                logger.warning('Identical serialized inputs with values (%f, %f)', As[i], Bs[j])
            else:
                iMatch = iMatch[0]
                output[i, j] = serialized_output[iMatch]
    return (As, Bs, output, nholes)


def rmse(x1, x2):
    """ Compute the root mean square error between two 1D arrays """
    return np.sqrt(((x1 - x2) ** 2).mean())


def rsquared(x1, x2):
    ''' compute the R-squared coefficient between two 1D arrays '''
    residuals = x1 - x2
    ss_res = np.sum(residuals**2)
    ss_tot = np.sum((x1 - np.mean(x1))**2)
    return 1 - (ss_res / ss_tot)


def DownSample(t_dense, y, nsparse):
    """ Decimate periodic signals to a specified number of samples."""

    if(y.ndim) > 1:
        nsignals = y.shape[0]
    else:
        nsignals = 1
        y = np.array([y])

    # determine time step and period of input signal
    T = t_dense[-1] - t_dense[0]
    dt_dense = t_dense[1] - t_dense[0]

    # resample time vector linearly
    t_ds = np.linspace(t_dense[0], t_dense[-1], nsparse)

    # create MAV window
    nmav = int(0.03 * T / dt_dense)
    if nmav % 2 == 0:
        nmav += 1
    mav = np.ones(nmav) / nmav

    # determine signals padding
    npad = int((nmav - 1) / 2)

    # determine indexes of sampling on convolved signals
    ids = np.round(np.linspace(0, t_dense.size - 1, nsparse)).astype(int)

    y_ds = np.empty((nsignals, nsparse))

    # loop through signals
    for i in range(nsignals):
        # pad, convolve and resample
        pad_left = y[i, -(npad + 2):-2]
        pad_right = y[i, 1:npad + 1]
        y_ext = np.concatenate((pad_left, y[i, :], pad_right), axis=0)
        y_mav = np.convolve(y_ext, mav, mode='valid')
        y_ds[i, :] = y_mav[ids]

    if nsignals == 1:
        y_ds = y_ds[0, :]

    return (t_ds, y_ds)


def Pressure2Intensity(p, rho=1075.0, c=1515.0):
    """ Return the spatial peak, pulse average acoustic intensity (ISPPA)
        associated with the specified pressure amplitude.

        :param p: pressure amplitude (Pa)
        :param rho: medium density (kg/m3)
        :param c: speed of sound in medium (m/s)
        :return: spatial peak, pulse average acoustic intensity (W/m2)
    """
    return p**2 / (2 * rho * c)


def Intensity2Pressure(I, rho=1075.0, c=1515.0):
    """ Return the pressure amplitude associated with the specified
        spatial peak, pulse average acoustic intensity (ISPPA).

        :param I: spatial peak, pulse average acoustic intensity (W/m2)
        :param rho: medium density (kg/m3)
        :param c: speed of sound in medium (m/s)
        :return: pressure amplitude (Pa)
    """
    return np.sqrt(2 * rho * c * I)


def find_nearest(array, value):
    ''' Find nearest element in 1D array. '''

    idx = (np.abs(array - value)).argmin()
    return (idx, array[idx])


def rescale(x, lb=None, ub=None, lb_new=0, ub_new=1):
    ''' Rescale a value to a specific interval by linear transformation. '''

    if lb is None:
        lb = x.min()
    if ub is None:
        ub = x.max()
    xnorm = (x - lb) / (ub - lb)
    return xnorm * (ub_new - lb_new) + lb_new


def LennardJones(x, beta, alpha, C, m, n):
    """ Generic expression of a Lennard-Jones function, adapted for the context of
        symmetric deflection (distance = 2x).

        :param x: deflection (i.e. half-distance)
        :param beta: x-shifting factor
        :param alpha: x-scaling factor
        :param C: y-scaling factor
        :param m: exponent of the repulsion term
        :param n: exponent of the attraction term
        :return: Lennard-Jones potential at given distance (2x)
    """
    return C * (np.power((alpha / (2 * x + beta)), m) - np.power((alpha / (2 * x + beta)), n))


def getNeuronsDict():
    ''' Return dictionary of neurons classes that can be instantiated. '''
    neurons_dict = {}
    for _, obj in inspect.getmembers(neurons):
        if inspect.isclass(obj) and isinstance(obj.name, str):
            neurons_dict[obj.name] = obj
    return neurons_dict


def get_BLS_lookups(a):
    lookup_path = getLookupDir() + '/BLS_lookups_a{:.1f}nm.json'.format(a * 1e9)
    try:
        with open(lookup_path) as fh:
            sample = json.load(fh)
        return sample
    except FileNotFoundError:
        return {}


def save_BLS_lookups(a, lookups):
    """ Save BLS parameter into specific lookup file
        :return: absolute path to the directory
    """
    lookup_path = getLookupDir() + '/BLS_lookups_a{:.1f}nm.json'.format(a * 1e9)
    with open(lookup_path, 'w') as fh:
        json.dump(lookups, fh)


def extractCompTimes(filenames):
    ''' Extract computation times from a list of simulation files. '''
    tcomps = np.empty(len(filenames))
    for i, fn in enumerate(filenames):
        logger.info('Loading data from "%s"', fn)
        with open(fn, 'rb') as fh:
            frame = pickle.load(fh)
            meta = frame['meta']
        tcomps[i] = meta['tcomp']
    return tcomps


def computeMeshEdges(x, scale='lin'):
    ''' Compute the appropriate edges of a mesh that quads a linear or logarihtmic distribution.

        :param x: the input vector
        :param scale: the type of distribution ('lin' for linear, 'log' for logarihtmic)
        :return: the edges vector
    '''

    if scale is 'log':
        x = np.log10(x)
    dx = x[1] - x[0]
    if scale is 'lin':
        y = np.linspace(x[0] - dx / 2, x[-1] + dx / 2, x.size + 1)
    elif scale is 'log':
        y = np.logspace(x[0] - dx / 2, x[-1] + dx / 2, x.size + 1)
    return y


si_prefixes = {
    'y': 1e-24,  # yocto
    'z': 1e-21,  # zepto
    'a': 1e-18,  # atto
    'f': 1e-15,  # femto
    'p': 1e-12,  # pico
    'n': 1e-9,   # nano
    'u': 1e-6,   # micro
    'm': 1e-3,   # mili
    '': 1e0,    # None
    'k': 1e3,    # kilo
    'M': 1e6,    # mega
    'G': 1e9,    # giga
    'T': 1e12,   # tera
    'P': 1e15,   # peta
    'E': 1e18,   # exa
    'Z': 1e21,   # zetta
    'Y': 1e24,   # yotta
}


def si_format(x, precision=0, space=''):
    ''' Format a float according to the SI unit system, with the appropriate prefix letter. '''
    if isinstance(x, float) or isinstance(x, int) or isinstance(x, np.float) or\
       isinstance(x, np.int32) or isinstance(x, np.int64):
        if x == 0:
            factor = 1e0
            prefix = ''
        else:
            sorted_si_prefixes = sorted(si_prefixes.items(), key=operator.itemgetter(1))
            vals = [tmp[1] for tmp in sorted_si_prefixes]
            # vals = list(si_prefixes.values())
            ix = np.searchsorted(vals, np.abs(x)) - 1
            if np.abs(x) == vals[ix + 1]:
                ix += 1
            factor = vals[ix]
            prefix = sorted_si_prefixes[ix][0]
            # prefix = list(si_prefixes.keys())[ix]
        return '{{:.{}f}}{}{}'.format(precision, space, prefix).format(x / factor)
    elif isinstance(x, list) or isinstance(x, tuple):
        return [si_format(item, precision, space) for item in x]
    elif isinstance(x, np.ndarray) and x.ndim == 1:
        return [si_format(float(item), precision, space) for item in x]
    else:
        print(type(x))


def getCycleAverage(t, y, T):
    ''' Compute the cycle-averaged profile of a signal given a specific periodicity.

        :param t: time vector (s)
        :param y: signal vector
        :param T: period (s)
        :return: cycle-averaged signal vector
    '''

    nsamples = y.size
    ncycles = int((t[-1] - t[0]) // T)
    npercycle = int(nsamples // ncycles)
    return np.mean(np.reshape(y[:ncycles * npercycle], (ncycles, npercycle)), axis=1)


def itrpLookupsFreq(lookups3D, freqs, Fdrive):
    """ Interpolate a dictionary of 3D lookups at a given frequency.

        :param lookups3D: dictionary of 3D lookups
        :param freqs: array of lookup frequencies (Hz)
        :param Fdrive: acoustic drive frequency (Hz)
        :return: a dictionary of 2D lookups interpolated a the given frequency
    """

    # If Fdrive in lookup frequencies, simply take (A, Q) slice at Fdrive index
    if Fdrive in freqs:
        iFdrive = np.searchsorted(freqs, Fdrive)
        # logger.debug('Using lookups directly at %.2f kHz', freqs[iFdrive] * 1e-3)
        lookups2D = {key: np.squeeze(lookups3D[key][iFdrive, :, :]) for key in lookups3D.keys()}

    # Otherwise, interpolate linearly between 2 (A, Q) slices at Fdrive bounding values indexes
    else:
        ilb = np.searchsorted(freqs, Fdrive) - 1
        iub = ilb + 1
        # logger.debug('Interpolating lookups between %.2f kHz and %.2f kHz',
        #              freqs[ilb] * 1e-3, freqs[iub] * 1e-3)
        lookups2D_lb = {key: np.squeeze(lookups3D[key][ilb, :, :]) for key in lookups3D.keys()}
        lookups2D_ub = {key: np.squeeze(lookups3D[key][iub, :, :]) for key in lookups3D.keys()}
        Fratio = (Fdrive - freqs[ilb]) / (freqs[iub] - freqs[ilb])
        lookups2D = {key: lookups2D_lb[key] + (lookups2D_ub[key] - lookups2D_lb[key]) * Fratio
                     for key in lookups3D.keys()}

    return lookups2D


def getLookups2D(mechname, a, Fdrive):
    ''' Retrieve appropriate 2D lookup tables and reference vectors
        for a given membrane mechanism, sonophore diameter and US frequency.

        :param mechname: name of membrane density mechanism
        :param a: sonophore diameter (m)
        :param Fdrive: US frequency (Hz)
        :return: 3-tuple with 1D numpy arrays of reference acoustic amplitudes and charge densities,
            and a dictionary of 2D lookup numpy arrays
    '''

    # Check lookup file existence
    lookup_file = '{}_lookups_a{:.1f}nm.pkl'.format(mechname, a * 1e9)
    lookup_path = '{}/{}'.format(getLookupDir(), lookup_file)
    if not os.path.isfile(lookup_path):
        raise InputError('Missing lookup file: "{}"'.format(lookup_file))

    # Load lookups dictionary
    with open(lookup_path, 'rb') as fh:
        lookups3D = pickle.load(fh)

    # Retrieve 1D inputs from lookups dictionary
    Fref = lookups3D.pop('f')
    Aref = lookups3D.pop('A')
    Qref = lookups3D.pop('Q')

    # Check that US frequency is within lookup range
    margin = 1e-9  # adding margin to compensate for eventual round error
    Frange = (Fref.min() - margin, Fref.max() + margin)
    if Fdrive < Frange[0] or Fdrive > Frange[1]:
        raise InputError('Invalid frequency: {}Hz (must be within {}Hz - {}Hz lookup interval)'
                         .format(*si_format([Fdrive, *Frange], precision=2, space=' ')))

    # Interpolate 3D lookups at US frequency
    lookups2D = itrpLookupsFreq(lookups3D, Fref, Fdrive)

    return Aref, Qref, lookups2D


def nDindexes(dims, index):
    ''' Find index positions in a n-dimensional array.

        :param dims: dimensions of the n-dimensional array (tuple or list)
        :param index: index position in the flattened n-dimensional array
        :return: list of indexes along each array dimension
    '''

    dims = list(dims)

    # Find size of each array dimension
    dims.reverse()
    dimsizes = [1]
    r = 1
    for x in dims[:-1]:
        r *= x
        dimsizes.append(r)
    dims.reverse()
    dimsizes.reverse()

    # Find indexes
    indexes = []
    remainder = index
    for dimsize in dimsizes[:-1]:
        idim, remainder = divmod(remainder, dimsize)
        indexes.append(idim)
    indexes.append(remainder)

    return indexes


def pow10_format(number, precision=2):
    ''' Format a number in power of 10 notation. '''
    ret_string = '{0:.{1:d}e}'.format(number, precision)
    a, b = ret_string.split("e")
    a = float(a)
    b = int(b)
    return '{}10^{{{}}}'.format('{} * '.format(a) if a != 1. else '', b)


def checkNumBounds(values, bounds):
    ''' Check if a set of numbers is within predefined bounds. '''

    # checking parameters against reference bounds
    for x, bound in zip(values, bounds):
        if x < bound[0] or x > bound[1]:
            raise ValueError('Input value {} out of [{}, {}] range'.format(x, bound[0], bound[1]))
    pass
